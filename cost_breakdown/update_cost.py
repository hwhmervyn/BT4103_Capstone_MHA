import csv
import openpyxl
import pytz
import datetime as dt
import pandas as pd
from enum import Enum

import sys, os
workingDirectory = os.getcwd()
costDirectory = os.path.join(workingDirectory, "cost_breakdown")

class Stage(Enum):
    EXCEL_FILTERING = "Excel Filtering"
    SUPPORT_ANALYSIS = "Support Analysis"
    IND_ANALYSIS = "Individual Analysis"
    AGG_ANALYSIS = "Aggregated Analysis"
    MISCELLANEOUS = "Miscellaneous"

singapore = pytz.timezone('Asia/Singapore')


## Excel Format
def update_usage_logs(processing_stage, query, total_input_tokens, total_output_tokens, total_cost):
    file_path = os.path.join(costDirectory, 'API Cost.xlsx')
    # Check if the file exists
    if not os.path.exists(file_path):
        # If the file doesn't exist, create an empty DataFrame
        columns = ['Date', 'Processing Stage', 'Query', 'Total Input Tokens', 'Total Output Tokens', 'Total Cost']
        df = pd.DataFrame(columns=columns)
        df.to_excel('cost_breakdown/API Cost.xlsx', sheet_name='Sheet1', index = False)
    
    workbook = openpyxl.load_workbook('cost_breakdown/API Cost.xlsx')
    worksheet = workbook['Sheet1']
    new_row_values = [str(dt.datetime.now()), processing_stage, query, total_input_tokens, total_output_tokens, total_cost]
    last_row = worksheet.max_row
    new_row = last_row + 1
    for col_num, value in enumerate(new_row_values, start=1):
        cell = worksheet.cell(row=new_row, column=col_num, value=value)
    workbook.save('cost_breakdown/API Cost.xlsx')

# WARNING! This deletes the rows in the .xlsx file
def clear_logs():
    # Delete Rows
    workbook = openpyxl.load_workbook('cost_breakdown/API Cost.xlsx')
    worksheet = workbook['Sheet1']
    max_row = worksheet.max_row
    for row_num in range(max_row, 0, -1):
        worksheet.delete_rows(row_num)
    
    #Add Headers
    new_row_values = ["Date","Processing Stage", "Query", "Total Input Tokens", "Total Output Tokens", "Total Cost"]
    last_row = worksheet.max_row
    new_row = last_row
    for col_num, value in enumerate(new_row_values, start=1):
        cell = worksheet.cell(row=new_row, column=col_num, value=value)
    workbook.save('cost_breakdown/API Cost.xlsx')


## CSV Format
'''
def update_usage_logs(processing_stage, query, total_input_tokens, total_output_tokens, total_cost):
    with open('cost_breakdown/API Cost.csv','a',newline="") as f_object:
        writer = csv.writer(f_object)
        writer.writerow([str(dt.datetime.now()), processing_stage, query, total_input_tokens, total_output_tokens, total_cost])
        f_object.close()
        
        
# WARNING! This deletes the rows in the .csv file
def clear_logs():
    f =  open('cost_breakdown/API Cost.csv','w')
    f.truncate()
    f.close()
    with open('cost_breakdown/API Cost.csv','a',newline="") as f_object:
        writer = csv.writer(f_object)
        writer.writerow(["Date","Processing Stage", "Query", "Total Input Tokens", "Total Output Tokens", "Total Cost"])
        f_object.close()
'''